import requests
import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import messagebox

# Function to get book details from Open Library
def get_book_details_open_library(isbn):
    url = f"https://openlibrary.org/api/books?bibkeys=ISBN:{isbn}&format=json&jscmd=data"
    response = requests.get(url)
    data = response.json()
    
    if f'ISBN:{isbn}' in data:
        book_data = data[f'ISBN:{isbn}']
        title = book_data.get('title', 'N/A')
        authors = book_data.get('authors', [])
        year = book_data.get('publish_date', 'N/A')
        
        return title, authors, year
    else:
        return None

# Function to get book details from Google Books API as a fallback
def get_book_details_google_books(isbn):
    url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
    response = requests.get(url)
    data = response.json()
    
    if 'items' in data:
        book_data = data['items'][0]['volumeInfo']
        title = book_data.get('title', 'N/A')
        authors = book_data.get('authors', ['N/A'])
        year = book_data.get('publishedDate', 'N/A')
        
        return title, authors, year
    else:
        return None

# Function to create custom reference ID
def create_reference_id(authors, year):
    if authors:
        author = authors[0]['name'] if isinstance(authors[0], dict) else authors[0]
        last_name, first_name = author.split(' ')[-1], author.split(' ')[0]
        reference_id = f"{last_name[:3].lower()}-{first_name[:3].lower()}-{year[-4:]}"  # Use last 4 digits of year
        return reference_id
    return 'N/A'

# Function to append or create Excel file
def update_excel(isbn, title, authors, year, reference_id):
    file_name = 'books.xlsx'
    existing_entries = set()

    try:
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            existing_entries.add(row[0])  # Assuming ISBN is in the first column
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(['ISBN', 'Book Title', 'Author(s)', 'Year', 'Reference ID'])

    if isbn not in existing_entries:
        sheet.append([isbn, title, authors, year, reference_id])
        wb.save(file_name)
        return True
    else:
        return False

# Function to handle ISBN input
def on_submit(event=None):  # Allow event parameter for Enter key
    isbn = isbn_entry.get()
    if isbn:
        # First attempt to get book details from Open Library
        book_details = get_book_details_open_library(isbn)
        if book_details:
            title, authors, year = book_details
        else:
            # Fallback to Google Books API
            book_details = get_book_details_google_books(isbn)
            if book_details:
                title, authors, year = book_details
            else:
                messagebox.showwarning("Error", "No book found for this ISBN.")
                return
        
        # Create custom reference ID
        reference_id = create_reference_id(authors, year)
        
        if update_excel(isbn, title, authors, year, reference_id):
            messagebox.showinfo("Success", f"Added:\nISBN: {isbn}\nTitle: {title}\nAuthors: {', '.join(authors)}\nYear: {year}\nReference ID: {reference_id}")
        else:
            messagebox.showwarning("Duplicate", "This book is already in the list.")
    else:
        messagebox.showwarning("Error", "Please enter an ISBN.")

# Create GUI
root = tk.Tk()
root.title("ISBN to Excel")

tk.Label(root, text="Enter ISBN:").pack(pady=10)
isbn_entry = tk.Entry(root, width=30)
isbn_entry.pack(pady=10)

# Bind the Enter key to the submit function
isbn_entry.bind('<Return>', on_submit)

submit_button = tk.Button(root, text="Submit", command=on_submit)
submit_button.pack(pady=10)

root.mainloop()


Exception in Tkinter callback
Traceback (most recent call last):
  File "/usr/lib/python3.10/tkinter/__init__.py", line 1921, in __call__
    return self.func(*args)
  File "<ipython-input-1-c42cbadec257>", line 90, in on_submit
    if update_excel(isbn, title, authors, year, reference_id):
  File "<ipython-input-1-c42cbadec257>", line 64, in update_excel
    sheet.append([isbn, title, authors, year, reference_id])
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/worksheet/worksheet.py", line 673, in append
    cell = Cell(self, row=row_idx, column=col_idx, value=content)
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/cell/cell.py", line 119, in __init__
    self.value = value
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/cell/cell.py", line 218, in value
    self._bind_value(value)
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/cell/cell.py", line 187, in _bind_value
    raise ValueError("Cannot convert {0!r} to Excel".format(value))
ValueError: Cannot convert ['Merlin Gideon Gray'] to Excel
Exception in Tkinter callback
Traceback (most recent call last):
  File "/usr/lib/python3.10/tkinter/__init__.py", line 1921, in __call__
    return self.func(*args)
  File "<ipython-input-1-c42cbadec257>", line 90, in on_submit
    if update_excel(isbn, title, authors, year, reference_id):
  File "<ipython-input-1-c42cbadec257>", line 64, in update_excel
    sheet.append([isbn, title, authors, year, reference_id])
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/worksheet/worksheet.py", line 673, in append
    cell = Cell(self, row=row_idx, column=col_idx, value=content)
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/cell/cell.py", line 119, in __init__
    self.value = value
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/cell/cell.py", line 218, in value
    self._bind_value(value)
  File "/home/gray221/.local/lib/python3.10/site-packages/openpyxl/cell/cell.py", line 187, in _bind_value
    raise ValueError("Cannot convert {0!r} to Excel".format(value))
ValueError: Cannot convert [{'url': 'https://openlibrary.org/authors/OL118077A/George_Orwell', 'name': 'George Orwell'}] to Excel